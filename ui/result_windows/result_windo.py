import tkinter as tk
from tkinter import ttk, filedialog
from tkinter import messagebox

import openpyxl
from openpyxl.utils import get_column_letter


class ResultWindow(tk.Toplevel):
    """설계 결과 요약 및 전주 상세보기"""

    def __init__(self, master, design_context):
        super().__init__(master)
        self.design_context = design_context

        self.title("설계 결과 요약")
        self.geometry("700x700")
        self.resizable(True, True)

        self.selected_pole = None

        self._build_ui()

    def _build_ui(self):
        container = ttk.Frame(self, padding=10)
        container.pack(fill="both", expand=True)

        # -----------------------------
        # 요약 Treeview
        # -----------------------------
        summary_frame = ttk.LabelFrame(container, text="전주 요약")
        summary_frame.pack(fill="both", expand=True, pady=(0, 10))

        columns = ("전주번호", "트랙", "측점", "X", "Y", "구간", "구조물", "선형")  # 예시로 열 추가
        self.tree_summary = ttk.Treeview(
            summary_frame,
            columns=columns,
            show="headings",
            selectmode="browse"
        )

        # 헤더 설정
        for col in columns:
            self.tree_summary.heading(col, text=col)
            self.tree_summary.column(col, width=100, anchor="center")

        # 세로 스크롤바
        vsb = ttk.Scrollbar(summary_frame, orient="vertical", command=self.tree_summary.yview)
        vsb.pack(side="right", fill="y")
        self.tree_summary.configure(yscrollcommand=vsb.set)

        # ✅ 가로 스크롤바
        hsb = ttk.Scrollbar(summary_frame, orient="horizontal", command=self.tree_summary.xview)
        hsb.pack(side="bottom", fill="x")
        self.tree_summary.configure(xscrollcommand=hsb.set)

        self.tree_summary.pack(fill="both", expand=True)

        # 전주 데이터 로드
        self._populate_summary()

        # 선택 이벤트
        self.tree_summary.bind("<<TreeviewSelect>>", self._on_pole_selected)

        # -----------------------------
        # 상세보기 Treeview
        # -----------------------------
        detail_frame = ttk.LabelFrame(container, text="선택 전주 상세")
        detail_frame.pack(fill="both", expand=True, pady=(0, 10))

        detail_columns = ("종류", "이름", "인덱스", "추가정보")
        self.tree_detail = ttk.Treeview(detail_frame, columns=detail_columns, show="headings")
        for col in detail_columns:
            self.tree_detail.heading(col, text=col)
            self.tree_detail.column(col, width=120, anchor="center")
        self.tree_detail.pack(fill="both", expand=True)

        # -----------------------------
        # 버튼 영역
        # -----------------------------
        btn_frame = ttk.Frame(container)
        btn_frame.pack(fill="x")

        ttk.Button(btn_frame, text="상세보기 초기화", command=self._clear_detail).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="출력", command=self._export).pack(side="right", padx=5)
        ttk.Button(btn_frame, text="닫기", command=self.destroy).pack(side="right")
    def _populate_summary(self):
        """요약 Treeview 데이터 채우기"""
        for pole in self.design_context.poledata.iter_poles():
            self.tree_summary.insert(
                "",
                "end",
                iid=id(pole),
                values=(
                    pole.post_number,
                    pole.track_index,
                    pole.pos,
                    getattr(pole.coord, "x", 0),
                    getattr(pole.coord, "y", 0),
                    getattr(pole.current_section, "name", ""),
                    pole.ref.structure_type,
                    pole.ref.curve_type
                )
            )

    def _on_pole_selected(self, event):
        """요약 Treeview에서 전주 선택 시 상세보기"""
        selected_id = self.tree_summary.selection()
        if not selected_id:
            return
        pole_obj_id = selected_id[0]

        # 선택 전주 객체
        self.selected_pole = None
        for pole in self.design_context.poledata.iter_poles():
            if id(pole) == int(pole_obj_id):
                self.selected_pole = pole
                break

        if self.selected_pole:
            self._populate_detail(self.selected_pole)

    def _populate_detail(self, pole):
        """상세보기 Treeview 데이터 채우기"""
        self._clear_detail()

        # Mast
        for mast in getattr(pole, "masts", []):
            self.tree_detail.insert("", "end", values=("Mast", getattr(mast, "name", ""), getattr(mast, "code", ""), ""))

        # Bracket
        for bracket in getattr(pole, "brackets", []):
            self.tree_detail.insert("", "end", values=("Bracket", getattr(bracket, "name", ""), getattr(bracket, "index", ""), ""))

        # Feeder
        for feeder in getattr(pole, "feeders", []):
            self.tree_detail.insert("", "end", values=("Feeder", getattr(feeder, "name", ""), getattr(feeder, "code", ""), ""))

    def _clear_detail(self):
        """상세보기 초기화"""
        for item in self.tree_detail.get_children():
            self.tree_detail.delete(item)

    def _export(self):
        """트리뷰 데이터를 Excel로 저장"""
        # 저장할 파일 경로 선택
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            title="엑셀 파일로 저장"
        )
        if not file_path:
            return  # 취소 시 종료

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pole Summary"

            # 컬럼 제목
            columns = ["Post Number", "Track Index", "Position", "Coord X", "Coord Y", "Section", "Structure", "Curve"]
            ws.append(columns)

            # 트리뷰 데이터 가져오기
            for iid in self.tree_summary.get_children():
                values = list(self.tree_summary.item(iid, "values"))
                # 숫자 컬럼 변환
                try:
                    values[1] = int(values[1])  # Track Index
                except ValueError:
                    pass
                try:
                    values[2] = float(values[2])  # Position
                    values[3] = float(values[3])  # Coord X
                    values[4] = float(values[4])  # Coord Y
                except ValueError:
                    pass
                ws.append(values)

            # 컬럼 너비 자동 조정
            for col_num, col in enumerate(columns, 1):
                max_length = max(
                    [len(str(ws.cell(row=row, column=col_num).value)) for row in range(1, ws.max_row + 1)]
                )
                ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

            wb.save(file_path)
            messagebox.showinfo("완료", f"엑셀 파일로 저장되었습니다.\n{file_path}")

        except Exception as e:
            messagebox.showerror("오류", f"엑셀 저장 중 오류 발생:\n{e}")

